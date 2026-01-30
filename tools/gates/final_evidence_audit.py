#!/usr/bin/env python3
"""
FINAL EVIDENCE AUDIT GATE

Verifies that every order marked Done in Excel has:
- Real evidence folder with latest run
- gate.log present
- verdict.json with end_to_end_working=true
- All linked features in FULL_STACK_FEATURE_MATRIX marked Is_End_to_End_Working="YES"

Exit 0 only if ALL checks pass. Otherwise exit non-zero with detailed report.
"""

import argparse
import json
import os
import sys
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Tuple

from openpyxl import load_workbook


DEBUG = False
MIN_GATE_LOG_BYTES = 2048  # Minimum gate.log size to prove real execution
EXECUTION_SIGNATURES = [
    "npm run",
    "flutter",
    "firebase",
    "jest",
    "playwright",
    "gradle",
    "BUILD SUCCESS",
    "BUILD FAILED",
    "PASS",
    "FAIL",
    "Tests:",
    "exit code",
    "Exit code"
]


def eprint(*args, **kwargs):
    """Print to stderr."""
    print(*args, file=sys.stderr, **kwargs)


def get_timestamp() -> str:
    """Get current timestamp in Beirut TZ."""
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S %Z")


def split_csv(val) -> List[str]:
    """Split CSV string into list."""
    if not val:
        return []
    return [x.strip() for x in str(val).split(",") if x and x.strip()]


def load_orders(wb) -> Tuple[List[Dict], str]:
    """Load ORDERS sheet. Returns (records, error)."""
    if "ORDERS" not in wb.sheetnames:
        return [], "Sheet 'ORDERS' not found"

    ws = wb["ORDERS"]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], "ORDERS sheet is empty"

    header = [str(c) if c is not None else "" for c in rows[0]]
    required = {"Order_ID", "Status", "Feature_IDs", "Evidence_Artifacts"}
    missing = required - set(header)
    if missing:
        return [], f"Missing columns: {', '.join(sorted(missing))}"

    records = []
    for r in rows[1:]:
        if all(c is None for c in r):
            continue
        obj = {}
        for h, v in zip(header, r):
            if h:
                obj[h] = v if v is not None else ""
        records.append(obj)

    return records, None


def load_feature_matrix(wb) -> Tuple[Dict[str, Dict], str]:
    """Load FULL_STACK_FEATURE_MATRIX. Returns (feature_map, error)."""
    if "FULL_STACK_FEATURE_MATRIX" not in wb.sheetnames:
        return {}, "Sheet 'FULL_STACK_FEATURE_MATRIX' not found"

    ws = wb["FULL_STACK_FEATURE_MATRIX"]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {}, "FULL_STACK_FEATURE_MATRIX is empty"

    header = [str(c) if c is not None else "" for c in rows[0]]
    required = {"Feature_ID", "Is_End_to_End_Working"}
    missing = required - set(header)
    if missing:
        return {}, f"Missing columns: {', '.join(sorted(missing))}"

    feature_map = {}
    for r in rows[1:]:
        if all(c is None for c in r):
            continue
        obj = {}
        for h, v in zip(header, r):
            if h:
                obj[h] = v if v is not None else ""
        fid = str(obj.get("Feature_ID", "")).strip()
        if fid:
            feature_map[fid] = obj

    return feature_map, None


def find_latest_evidence(evidence_root: Path, order_id: str) -> Tuple[Path, str]:
    """
    Find latest evidence folder for order.
    Returns (evidence_path, error).
    """
    order_evidence_dir = evidence_root / order_id
    if not order_evidence_dir.exists():
        return None, f"No evidence directory for {order_id}"

    # Get all run folders (format: YYYYMMDD-HHMMSS)
    run_folders = [d for d in order_evidence_dir.iterdir() if d.is_dir()]
    if not run_folders:
        return None, f"No run folders in {order_evidence_dir}"

    # Sort lexicographically (YYYYMMDD-HHMMSS format sorts correctly)
    run_folders.sort(reverse=True)
    latest = run_folders[0]

    return latest, None


def validate_evidence_folder(evidence_path: Path, order_gate_command: str) -> List[str]:
    """
    Validate evidence folder contents with strict execution checks.
    Returns list of errors (empty if valid).
    """
    errors = []

    # Check gate.log exists
    gate_log = evidence_path / "gate.log"
    if not gate_log.exists():
        errors.append(f"Missing gate.log in {evidence_path}")
    else:
        # Check minimum size
        size = gate_log.stat().st_size
        if size < MIN_GATE_LOG_BYTES:
            errors.append(f"gate.log too small ({size} bytes < {MIN_GATE_LOG_BYTES} bytes minimum) in {evidence_path}")
        
        # Check for execution signatures
        try:
            with open(gate_log, 'r', errors='replace') as f:
                log_content = f.read()
            
            signature_count = sum(1 for sig in EXECUTION_SIGNATURES if sig.lower() in log_content.lower())
            if signature_count < 2:
                errors.append(f"gate.log shows insufficient execution signatures ({signature_count} < 2 required) in {evidence_path}")
        except Exception as e:
            errors.append(f"Failed to read gate.log in {evidence_path}: {e}")

    # Check verdict.json exists and is valid
    verdict_path = evidence_path / "verdict.json"
    if not verdict_path.exists():
        errors.append(f"Missing verdict.json in {evidence_path}")
    else:
        try:
            with open(verdict_path, 'r') as f:
                verdict = json.load(f)
            
            # Validate structure
            if not isinstance(verdict, dict):
                errors.append(f"verdict.json is not a JSON object in {evidence_path}")
            else:
                # Check required fields
                if "order_id" not in verdict:
                    errors.append(f"verdict.json missing 'order_id' field in {evidence_path}")
                
                if "run_ts" not in verdict:
                    errors.append(f"verdict.json missing 'run_ts' field in {evidence_path}")
                
                if "end_to_end_working" not in verdict:
                    errors.append(f"verdict.json missing 'end_to_end_working' field in {evidence_path}")
                elif verdict.get("end_to_end_working") is not True:
                    errors.append(f"verdict.json has end_to_end_working={verdict.get('end_to_end_working')} (expected true) in {evidence_path}")
                
                # Check artifacts list
                artifacts = verdict.get("artifacts", [])
                if not isinstance(artifacts, list) or len(artifacts) < 1:
                    errors.append(f"verdict.json missing or empty 'artifacts' list in {evidence_path}")
                
                # Check mtime ordering: verdict created after gate.log
                if gate_log.exists():
                    verdict_mtime = verdict_path.stat().st_mtime
                    gate_log_mtime = gate_log.stat().st_mtime
                    if verdict_mtime < gate_log_mtime:
                        errors.append(f"verdict.json created BEFORE gate.log (not derived from execution) in {evidence_path}")
        
        except json.JSONDecodeError as e:
            errors.append(f"verdict.json is invalid JSON in {evidence_path}: {e}")
        except Exception as e:
            errors.append(f"Failed to read verdict.json in {evidence_path}: {e}")

    return errors


def audit_orders(
    orders: List[Dict],
    feature_map: Dict[str, Dict],
    evidence_root: Path
) -> Tuple[List[Dict], bool]:
    """
    Audit all Done orders with strict validation.
    Returns (audit_results, overall_pass).
    """
    results = []
    overall_pass = True

    for order in orders:
        order_id = str(order.get("Order_ID", "")).strip()
        status = str(order.get("Status", "")).strip()
        gate_command = str(order.get("Gate_Command", "")).strip()

        if status != "Done":
            continue  # Only audit Done orders

        result = {
            "order_id": order_id,
            "status": status,
            "gate_command": gate_command,
            "passed": True,
            "errors": []
        }

        # Check gate_command is non-trivial
        if not gate_command:
            result["errors"].append(f"Order {order_id} has empty Gate_Command")
            result["passed"] = False
        elif gate_command.startswith("echo ") and len(gate_command) < 50:
            result["errors"].append(f"Order {order_id} has trivial Gate_Command (echo-only): {gate_command}")
            result["passed"] = False

        # Check evidence folder
        evidence_path, err = find_latest_evidence(evidence_root, order_id)
        if err:
            result["errors"].append(err)
            result["passed"] = False
        else:
            result["evidence_path"] = str(evidence_path.relative_to(evidence_root.parent))
            
            # Validate evidence folder contents with strict checks
            validation_errors = validate_evidence_folder(evidence_path, gate_command)
            result["errors"].extend(validation_errors)
            if validation_errors:
                result["passed"] = False

        # Check linked features
        feature_ids = split_csv(order.get("Feature_IDs", ""))
        if feature_ids:
            feature_errors = []
            for fid in feature_ids:
                if fid not in feature_map:
                    feature_errors.append(f"Feature {fid} not found in FULL_STACK_FEATURE_MATRIX")
                else:
                    e2e = str(feature_map[fid].get("Is_End_to_End_Working", "")).upper().strip()
                    if e2e != "YES":
                        feature_errors.append(f"Feature {fid} has Is_End_to_End_Working='{e2e}' (expected 'YES')")
            
            if feature_errors:
                result["errors"].extend(feature_errors)
                result["passed"] = False

        if not result["passed"]:
            overall_pass = False

        results.append(result)

    return results, overall_pass


def generate_audit_md(results: List[Dict], overall_pass: bool, timestamp: str) -> str:
    """Generate human-readable audit report."""
    lines = [
        "# Final Evidence Audit Report",
        f"**Timestamp:** {timestamp}",
        f"**Overall Status:** {'✅ PASS' if overall_pass else '❌ FAIL'}",
        "",
        "## Summary",
        ""
    ]

    if overall_pass:
        lines.append(f"All {len(results)} Done orders have valid evidence and feature linkage.")
        lines.append("")
        lines.append("**Decision:** Project completion verified. All claims backed by evidence.")
    else:
        failed = [r for r in results if not r["passed"]]
        lines.append(f"❌ {len(failed)} of {len(results)} Done orders have issues:")
        lines.append("")
        
        for result in failed:
            lines.append(f"### {result['order_id']}")
            lines.append("**Errors:**")
            for err in result["errors"]:
                lines.append(f"- {err}")
            lines.append("")
        
        lines.append("**Decision:** Project NOT complete. Orders marked Done without proper evidence.")
        lines.append("")
        lines.append("## Remediation")
        lines.append("1. Review each failed order above")
        lines.append("2. Re-run gates to generate proper evidence")
        lines.append("3. Verify features are truly e2e working")
        lines.append("4. Update Excel statuses to reflect reality")

    return "\n".join(lines)


def main():
    global DEBUG
    ap = argparse.ArgumentParser(description="Final evidence audit gate")
    ap.add_argument("--excel", default="UrbanPoints_CTO_Master_Control_v4.xlsx",
                    help="Path to Excel control file")
    ap.add_argument("--evidence-root", default="docs/evidence",
                    help="Root directory for evidence")
    ap.add_argument("--output-dir", help="Output directory for audit results")
    ap.add_argument("--debug", action="store_true", help="Enable debug output")
    args = ap.parse_args()

    DEBUG = args.debug

    # Paths
    excel_path = Path(args.excel).expanduser().resolve()
    evidence_root = Path(args.evidence_root).expanduser().resolve()

    if DEBUG:
        eprint(f"[DEBUG] Excel: {excel_path}")
        eprint(f"[DEBUG] Evidence root: {evidence_root}")

    # Load Excel
    if not excel_path.exists():
        eprint(f"ERROR: Excel file not found: {excel_path}")
        sys.exit(1)

    try:
        wb = load_workbook(excel_path, data_only=True)
    except Exception as e:
        eprint(f"ERROR: Failed to load Excel: {e}")
        sys.exit(1)

    # Load sheets
    orders, err = load_orders(wb)
    if err:
        eprint(f"ERROR: {err}")
        sys.exit(1)

    feature_map, err = load_feature_matrix(wb)
    if err:
        eprint(f"ERROR: {err}")
        sys.exit(1)

    if DEBUG:
        eprint(f"[DEBUG] Loaded {len(orders)} orders")
        eprint(f"[DEBUG] Loaded {len(feature_map)} features")

    # Audit
    timestamp = get_timestamp()
    results, overall_pass = audit_orders(orders, feature_map, evidence_root)

    if DEBUG:
        eprint(f"[DEBUG] Audited {len(results)} Done orders")
        eprint(f"[DEBUG] Overall pass: {overall_pass}")

    # Generate outputs
    audit_json = {
        "timestamp": timestamp,
        "overall_pass": overall_pass,
        "total_done_orders": len(results),
        "passed_orders": len([r for r in results if r["passed"]]),
        "failed_orders": len([r for r in results if not r["passed"]]),
        "results": results
    }

    audit_md = generate_audit_md(results, overall_pass, timestamp)

    # Write outputs
    if args.output_dir:
        output_dir = Path(args.output_dir)
        output_dir.mkdir(parents=True, exist_ok=True)
        
        with open(output_dir / "audit.json", 'w') as f:
            json.dump(audit_json, f, indent=2)
        
        with open(output_dir / "audit.md", 'w') as f:
            f.write(audit_md)
        
        if DEBUG:
            eprint(f"[DEBUG] Wrote audit.json and audit.md to {output_dir}")
    else:
        # Print to stdout
        print(json.dumps(audit_json, indent=2))
        print("\n" + audit_md)

    # Exit code
    sys.exit(0 if overall_pass else 1)


if __name__ == "__main__":
    main()
