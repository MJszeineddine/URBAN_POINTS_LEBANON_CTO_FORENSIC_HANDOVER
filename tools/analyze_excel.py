import sys
import json
from pathlib import Path
from datetime import datetime
from typing import Any, Dict, List

from openpyxl import load_workbook

CANONICAL = {
    "backend": {"percent": 75, "status": "PARTIAL"},
    "webadmin": {"percent": 85, "status": "PARTIAL"},
    "mobilecustomer": {"percent": 85, "status": "PARTIAL"},
    "mobilemerchant": {"percent": 70, "status": "PARTIAL"},
    "automation": {"percent": 50, "status": "PARTIAL"},
    "deployment": {"percent": 0, "status": "CRITICAL"},
    "documentation": {"percent": 85, "status": "GOOD"},
    "qa": {"percent": 60, "status": "PARTIAL"},
}

ALIASES = {
    "web admin": "webadmin",
    "web-admin": "webadmin",
    "webadmin": "webadmin",
    "mobile customer": "mobilecustomer",
    "mobile-customer": "mobilecustomer",
    "mobilecustomer": "mobilecustomer",
    "mobile merchant": "mobilemerchant",
    "mobile-merchant": "mobilemerchant",
    "mobilemerchant": "mobilemerchant",
    "devops": "deployment",
    "deployment": "deployment",
    "docs": "documentation",
    "documentation": "documentation",
}

STATUS_NORMALIZE = {
    "done": "DONE",
    "complete": "DONE",
    "working": "DONE",
    "partial": "PARTIAL",
    "in progress": "PARTIAL",
    "implemented": "PARTIAL",
    "implemented-not-active": "PARTIAL",
    "disabled": "DISABLED",
    "not done": "NOT DONE",
    "missing": "NOT DONE",
    "critical": "CRITICAL",
}


def slug(s: str) -> str:
    return ''.join(ch.lower() for ch in s if ch.isalnum())


def normalize_component(name: str) -> str:
    key = slug(name)
    # direct canonical match
    if key in CANONICAL:
        return key
    # alias match by token form
    for alias, target in ALIASES.items():
        if slug(alias) == key:
            return target
    # heuristic for backend, qa, automation, etc.
    if "backend" in key:
        return "backend"
    if "web" in key and "admin" in key:
        return "webadmin"
    if "customer" in key:
        return "mobilecustomer"
    if "merchant" in key:
        return "mobilemerchant"
    if "auto" in key or "scheduler" in key:
        return "automation"
    if "deploy" in key or "devops" in key:
        return "deployment"
    if "doc" in key:
        return "documentation"
    if key == "qa" or "quality" in key:
        return "qa"
    return key


def normalize_status(val: Any) -> str:
    if val is None:
        return ""
    s = str(val).strip()
    low = s.lower()
    for k, v in STATUS_NORMALIZE.items():
        if k in low:
            return v
    # if looks like a percent
    try:
        n = float(str(s).replace('%','').strip())
        if n >= 95:
            return "DONE"
        if n >= 70:
            return "PARTIAL"
        if n > 0:
            return "PARTIAL"
        return "NOT DONE"
    except Exception:
        pass
    return s


def extract_rows(ws) -> List[Dict[str, Any]]:
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    # find header row: first non-empty row
    header = None
    idx = 0
    for i, r in enumerate(rows):
        if any(cell is not None and str(cell).strip() for cell in r):
            header = [str(c).strip() if c is not None else '' for c in r]
            idx = i + 1
            break
    if header is None:
        return []
    data = []
    for r in rows[idx:]:
        if all(c is None or str(c).strip()=='' for c in r):
            continue
        row = {}
        for h, v in zip(header, r):
            if h:
                row[h] = v
        data.append(row)
    return data


def summarize_components(parsed: Dict[str, List[Dict[str, Any]]]) -> Dict[str, Dict[str, Any]]:
    summary: Dict[str, Dict[str, Any]] = {}
    # heuristic: look for columns named like 'Component', 'Area', 'Capability', 'Status', '%', 'Percent'
    for sheet, rows in parsed.items():
        for row in rows:
            headers = {slug(h): h for h in row.keys()}
            comp_key = None
            for candidate in ["component", "area", "capability", "module", "name"]:
                if candidate in headers:
                    comp_key = headers[candidate]
                    break
            if not comp_key:
                continue
            comp_raw = str(row.get(comp_key, '')).strip()
            if not comp_raw:
                continue
            comp = normalize_component(comp_raw)
            # try to get status
            status_val = None
            for candidate in ["status", "backendstatus", "frontendstatus", "automationstatus", "state"]:
                if candidate in headers:
                    status_val = row.get(headers[candidate])
                    if status_val:
                        break
            percent_val = None
            for candidate in ["%", "percent", "completion", "progress"]:
                key = candidate
                if key in headers:
                    percent_val = row.get(headers[key])
                    break
            status_norm = normalize_status(status_val)
            try:
                percent_num = None
                if percent_val is not None and str(percent_val).strip() != "":
                    percent_num = float(str(percent_val).replace('%','').strip())
            except Exception:
                percent_num = None
            # prefer highest certainty: percent > status text
            entry = summary.setdefault(comp, {"statuses": [], "percents": []})
            if status_norm:
                entry["statuses"].append(status_norm)
            if percent_num is not None:
                entry["percents"].append(percent_num)
    # collapse
    collapsed: Dict[str, Dict[str, Any]] = {}
    for comp, v in summary.items():
        percents = v.get("percents", [])
        statuses = v.get("statuses", [])
        avg_percent = round(sum(percents)/len(percents), 1) if percents else None
        # choose most frequent status if present
        status_final = None
        if statuses:
            from collections import Counter
            status_final = Counter(statuses).most_common(1)[0][0]
        collapsed[comp] = {
            "percent": avg_percent,
            "status": status_final,
            "raw": v
        }
    return collapsed


def compare_with_canonical(excel_summary: Dict[str, Dict[str, Any]]):
    comparisons: List[Dict[str, Any]] = []
    for key, canon in CANONICAL.items():
        excel = excel_summary.get(key)
        comp = {"component": key, "canonical": canon, "excel": excel}
        # compute mismatch flags
        status_mismatch = None
        percent_mismatch = None
        if excel:
            s_exc = (excel.get("status") or "").upper()
            s_can = canon["status"].upper()
            if s_exc and s_exc != s_can:
                status_mismatch = {"excel": s_exc, "canonical": s_can}
            p_exc = excel.get("percent")
            if p_exc is not None and abs(p_exc - canon["percent"]) >= 10:
                percent_mismatch = {"excel": p_exc, "canonical": canon["percent"]}
        else:
            status_mismatch = {"excel": None, "canonical": canon["status"]}
            percent_mismatch = {"excel": None, "canonical": canon["percent"]}
        comp["status_mismatch"] = status_mismatch
        comp["percent_mismatch"] = percent_mismatch
        comparisons.append(comp)
    return comparisons


def main():
    if len(sys.argv) < 2:
        print("Usage: analyze_excel.py <path_to_xlsx>")
        sys.exit(2)
    excel_path = Path(sys.argv[1]).expanduser().resolve()
    if not excel_path.exists():
        print(f"Excel not found: {excel_path}")
        sys.exit(2)

    wb = load_workbook(excel_path, data_only=True)
    parsed: Dict[str, List[Dict[str, Any]]] = {}
    for ws in wb.worksheets:
        parsed[ws.title] = extract_rows(ws)

    # outputs
    ts = datetime.utcnow().strftime('%Y%m%dT%H%M%SZ')
    out_dir = Path('docs/evidence/excel_audit') / ts
    out_dir.mkdir(parents=True, exist_ok=True)

    parsed_file = out_dir / 'parsed.json'
    with parsed_file.open('w', encoding='utf-8') as f:
        json.dump(parsed, f, ensure_ascii=False, indent=2, default=str)

    excel_summary = summarize_components(parsed)
    summary_file = out_dir / 'excel_summary.json'
    with summary_file.open('w', encoding='utf-8') as f:
        json.dump(excel_summary, f, ensure_ascii=False, indent=2)

    comparisons = compare_with_canonical(excel_summary)
    mismatches = [c for c in comparisons if c["status_mismatch"] or c["percent_mismatch"]]

    report_file = out_dir / 'mismatch_report.md'
    with report_file.open('w', encoding='utf-8') as f:
        f.write('# Excel vs CTO Report — Mismatch Report\n\n')
        f.write(f'Excel: {excel_path.name}\n\n')
        for comp in comparisons:
            name = comp['component']
            canon = comp['canonical']
            excel = comp['excel']
            f.write(f'## {name}\n')
            f.write(f'- Canonical: status={canon["status"]}, percent={canon["percent"]}\n')
            if excel:
                f.write(f'- Excel: status={excel.get("status")}, percent={excel.get("percent")}\n')
            else:
                f.write('- Excel: not found\n')
            if comp['status_mismatch']:
                f.write(f'- Status mismatch: {comp["status_mismatch"]}\n')
            if comp['percent_mismatch']:
                f.write(f'- Percent mismatch: {comp["percent_mismatch"]}\n')
            f.write('\n')

    # console summary
    print('Excel parsed. Outputs:')
    print(f'- {parsed_file}')
    print(f'- {summary_file}')
    print(f'- {report_file}')

if __name__ == '__main__':
    main()
