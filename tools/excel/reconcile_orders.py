#!/usr/bin/env python
"""
AUTO-RECONCILE: Synchronize Order Status with FULL_STACK_FEATURE_MATRIX reality.

When the autonomous loop stalls (no_open_orders), this script:
1. Reads FULL_STACK_FEATURE_MATRIX to find features NOT e2e working (Is_End_to_End_Working != 'YES')
2. Finds orders that depend on those features
3. Sets order Status to 'Open' (if deps satisfied) or 'Blocked' (if deps not satisfied)
4. Logs changes to CHANGELOG
5. Outputs reconciliation statistics

Usage:
  python3 tools/excel/reconcile_orders.py [--excel FILE]
  
Environment:
  EXCEL_PATH: Override Excel file path
  TZ: Timezone (default Asia/Beirut)
"""

import argparse
import json
import os
import sys
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Set, Tuple

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException


DEBUG = False


def eprint(*args, **kwargs):
    """Print to stderr."""
    print(*args, file=sys.stderr, **kwargs)


def output_json(ok: bool, **kwargs):
    """Output JSON response to stdout and exit."""
    response = {"ok": ok}
    response.update(kwargs)
    print(json.dumps(response))
    sys.exit(0)


def get_timestamp() -> str:
    """Get current timestamp in Beirut TZ."""
    # Respect TZ env var for test isolation
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S %Z")


def load_feature_matrix(wb) -> Tuple[Dict[str, Dict], Optional[str]]:
    """Load FULL_STACK_FEATURE_MATRIX. Returns (feature_map, error_message)."""
    if "FULL_STACK_FEATURE_MATRIX" not in wb.sheetnames:
        return {}, "Sheet 'FULL_STACK_FEATURE_MATRIX' not found"

    try:
        ws = wb["FULL_STACK_FEATURE_MATRIX"]
    except Exception as e:
        return {}, f"Failed to read FULL_STACK_FEATURE_MATRIX: {e}"

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {}, "FULL_STACK_FEATURE_MATRIX is empty"

    header = [str(c) if c is not None else "" for c in rows[0]]
    if DEBUG:
        eprint(f"[DEBUG] FSFM header: {header}")

    required_cols = {"Feature_ID", "Is_End_to_End_Working"}
    missing = required_cols - set(header)
    if missing:
        return {}, f"Missing columns in FULL_STACK_FEATURE_MATRIX: {', '.join(sorted(missing))}"

    feature_map = {}
    for r in rows[1:]:
        if all(c is None for c in r):
            continue
        obj = {}
        for h, v in zip(header, r):
            if h:
                obj[h] = v if v is not None else ""
        feature_id = str(obj.get("Feature_ID", "")).strip()
        if feature_id:
            feature_map[feature_id] = obj

    if DEBUG:
        eprint(f"[DEBUG] Loaded {len(feature_map)} features from matrix")

    return feature_map, None


def load_orders(wb) -> Tuple[List[Dict], List[int], Optional[str]]:
    """Load ORDERS sheet. Returns (records, row_indices, error_message)."""
    if "ORDERS" not in wb.sheetnames:
        return [], [], "Sheet 'ORDERS' not found"

    try:
        ws = wb["ORDERS"]
    except Exception as e:
        return [], [], f"Failed to read ORDERS sheet: {e}"

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], [], "ORDERS sheet is empty"

    header = [str(c) if c is not None else "" for c in rows[0]]
    required_cols = {"Order_ID", "Status", "Feature_IDs", "Depends_On_Orders", "Priority"}
    missing = required_cols - set(header)
    if missing:
        return [], [], f"Missing columns in ORDERS: {', '.join(sorted(missing))}"

    if DEBUG:
        eprint(f"[DEBUG] ORDERS header: {header}")

    records = []
    row_indices = []
    for i, r in enumerate(rows[1:], start=1):
        if all(c is None for c in r):
            continue
        obj = {}
        for h, v in zip(header, r):
            if h:
                obj[h] = v if v is not None else ""
        records.append(obj)
        row_indices.append(i)

    if DEBUG:
        eprint(f"[DEBUG] Loaded {len(records)} order records")

    return records, row_indices, None


def load_changelog(wb) -> Tuple[List[Dict], Optional[str]]:
    """Load CHANGELOG sheet. Returns (records, error_message)."""
    if "CHANGELOG" not in wb.sheetnames:
        # CHANGELOG doesn't exist yet; we'll create it when appending
        return [], None

    try:
        ws = wb["CHANGELOG"]
    except Exception as e:
        return [], f"Failed to read CHANGELOG: {e}"

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], None

    header = [str(c) if c is not None else "" for c in rows[0]]
    records = []
    for r in rows[1:]:
        if all(c is None for c in r):
            continue
        obj = {}
        for h, v in zip(header, r):
            if h:
                obj[h] = v if v is not None else ""
        records.append(obj)

    return records, None


def split_csv(val) -> List[str]:
    """Split CSV string into list."""
    if not val:
        return []
    return [x.strip() for x in str(val).split(",") if x and x.strip()]


def reconcile_orders(
    feature_map: Dict[str, Dict],
    records: List[Dict],
    row_indices: List[int]
) -> Tuple[List[Tuple[int, str, str, str]], Optional[str]]:
    """
    Reconcile order statuses.
    
    Returns (changes, error_message) where changes is:
      [(row_idx, order_id, old_status, new_status), ...]
    """
    changes = []
    status_map = {str(rec.get("Order_ID", "")): str(rec.get("Status", "")) for rec in records}

    # Find features that are NOT e2e working
    not_working_features = {
        fid for fid, fdata in feature_map.items()
        if str(fdata.get("Is_End_to_End_Working", "")).upper() != "YES"
    }

    if DEBUG:
        eprint(f"[DEBUG] Features NOT e2e working: {not_working_features}")

    # For each order: determine if it should be Open or Blocked
    for i, (rec, row_idx) in enumerate(zip(records, row_indices)):
        order_id = str(rec.get("Order_ID", "")).strip()
        current_status = str(rec.get("Status", "")).strip()
        feature_ids = split_csv(rec.get("Feature_IDs", ""))
        depends_on = split_csv(rec.get("Depends_On_Orders", ""))

        # Determine if this order's features are working
        order_features_working = all(
            fid not in not_working_features for fid in feature_ids
        ) if feature_ids else True

        # Determine if dependencies are satisfied
        deps_satisfied = all(
            status_map.get(dep, "") == "Done" for dep in depends_on
        ) if depends_on else True

        if DEBUG:
            eprint(f"[DEBUG] {order_id}: features_working={order_features_working}, "
                   f"deps_satisfied={deps_satisfied}, current_status={current_status}")

        # Reconcile logic:
        # - If features are NOT working AND order is Done -> reopen it (Open/Blocked based on deps)
        # - If features ARE working AND order is Done -> LEAVE IT (it's legitimately complete)
        # - If features are NOT working AND order is Open/Blocked -> keep it Open/Blocked based on deps
        target_status = ""
        
        if not order_features_working:
            # Features not yet done; order should NOT be Done
            if current_status == "Done":
                # This is wrong - features aren't working but order marked Done
                target_status = "Open" if deps_satisfied else "Blocked"
            elif current_status in ("Open", "Blocked", "In Progress"):
                # Already in progress/blocked state; adjust based on deps
                new_state = "Open" if deps_satisfied else "Blocked"
                if current_status != new_state:
                    target_status = new_state
        else:
            # Features ARE working
            if current_status == "Done":
                # This is correct - features working and order Done. LEAVE IT ALONE.
                pass
            elif current_status in ("Open", "Blocked", "In Progress"):
                # Order not Done but features are working - could mean it needs to run
                # Only change Blocked->Open if deps now satisfied
                if current_status == "Blocked" and deps_satisfied:
                    target_status = "Open"

        # Only log change if status differs
        if target_status and current_status != target_status:
            changes.append((row_idx, order_id, current_status, target_status))
            status_map[order_id] = target_status  # Update for next iterations

    return changes, None


def apply_changes(wb, changes: List[Tuple[int, str, str, str]], evidence_ts: str):
    """Apply status changes to ORDERS sheet and append CHANGELOG."""
    if not changes:
        return True

    # Update ORDERS sheet
    ws = wb["ORDERS"]
    # Get header from first row
    header_row = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
    header = [str(c) if c is not None else "" for c in header_row]
    
    try:
        status_col_idx = header.index("Status") + 1
    except ValueError:
        eprint(f"ERROR: 'Status' column not found in ORDERS. Header: {header}")
        return False
    
    for row_idx, order_id, old_status, new_status in changes:
        cell = ws.cell(row=row_idx + 1, column=status_col_idx)
        cell.value = new_status
        if DEBUG:
            eprint(f"[DEBUG] Updated {order_id} row {row_idx + 1}: {old_status} -> {new_status}")

    # Append to CHANGELOG if it exists
    if "CHANGELOG" in wb.sheetnames:
        ws_cl = wb["CHANGELOG"]
        changelog_rows = list(ws_cl.iter_rows(values_only=True))
        next_row = len(changelog_rows) + 1

        for row_idx, order_id, old_status, new_status in changes:
            change_id = f"AUTO-{evidence_ts.replace(' ', '-').replace(':', '')}"
            ws_cl.cell(row=next_row, column=1).value = change_id
            ws_cl.cell(row=next_row, column=2).value = evidence_ts
            ws_cl.cell(row=next_row, column=3).value = "ORDERS"
            ws_cl.cell(row=next_row, column=4).value = order_id  # Row_Key
            ws_cl.cell(row=next_row, column=5).value = "Status"  # Field
            ws_cl.cell(row=next_row, column=6).value = old_status  # Old_Value
            ws_cl.cell(row=next_row, column=7).value = new_status  # New_Value
            ws_cl.cell(row=next_row, column=8).value = "auto-reconcile from FULL_STACK_FEATURE_MATRIX"
            ws_cl.cell(row=next_row, column=9).value = "excel:FULL_STACK_FEATURE_MATRIX"
            next_row += 1

    return True


def main():
    global DEBUG
    ap = argparse.ArgumentParser(
        description="Auto-reconcile order statuses with feature matrix"
    )
    ap.add_argument("--excel", default=None, help="Path to Excel control file")
    ap.add_argument("--debug", action="store_true", help="Enable debug output to stderr")
    args = ap.parse_args()

    DEBUG = args.debug

    # Determine Excel path
    excel_path = args.excel or os.getenv("EXCEL_PATH", "UrbanPoints_CTO_Master_Control_v4.xlsx")
    path = Path(excel_path).expanduser().resolve()

    if DEBUG:
        eprint(f"[DEBUG] Excel path: {path}")
        eprint(f"[DEBUG] Excel exists: {path.exists()}")

    # Load workbook
    if not path.exists():
        output_json(False, error="excel_not_found", message=f"Excel file not found: {path}")

    try:
        wb = load_workbook(path)
    except InvalidFileException as e:
        output_json(False, error="invalid_excel", message=f"Invalid Excel file: {e}")
    except Exception as e:
        output_json(False, error="load_excel_failed", message=f"Failed to load Excel: {e}")

    # Load sheets
    feature_map, err = load_feature_matrix(wb)
    if err:
        output_json(False, error="load_feature_matrix_failed", message=err)

    records, row_indices, err = load_orders(wb)
    if err:
        output_json(False, error="load_orders_failed", message=err)

    # Load existing changelog (not critical if it fails)
    changelog_records, _ = load_changelog(wb)

    # Reconcile
    changes, err = reconcile_orders(feature_map, records, row_indices)
    if err:
        output_json(False, error="reconcile_failed", message=err)

    if DEBUG:
        eprint(f"[DEBUG] Identified {len(changes)} changes to apply")

    # Get timestamp for evidence
    ts = get_timestamp()

    # Apply changes
    if not apply_changes(wb, changes, ts):
        output_json(False, error="apply_changes_failed", message="Failed to apply changes to Excel")

    # Save workbook
    try:
        wb.save(path)
    except Exception as e:
        output_json(False, error="save_excel_failed", message=f"Failed to save Excel: {e}")

    # Determine which order to flag as the "reopened" order
    reopened_order_id = None
    changed_order_ids = [oid for _, oid, _, new_status in changes if new_status == "Open"]

    # Pick the highest priority Open order from the changes
    if changed_order_ids:
        PRIORITY_ORDER = {"P0": 0, "P1": 1, "P2": 2}
        best_priority = 99
        for rec in records:
            oid = rec.get("Order_ID")
            if oid in changed_order_ids:
                pr = str(rec.get("Priority", "P2"))
                pr_val = PRIORITY_ORDER.get(pr, 99)
                if pr_val < best_priority:
                    best_priority = pr_val
                    reopened_order_id = oid

    # Output success
    output_json(
        True,
        action="reconcile",
        orders_changed=len(changes),
        reopened_order_id=reopened_order_id,
        changed_order_ids=[oid for _, oid, _, _ in changes],
        timestamp=ts
    )


if __name__ == "__main__":
    main()
