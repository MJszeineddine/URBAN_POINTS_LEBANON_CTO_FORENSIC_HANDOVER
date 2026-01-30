#!/usr/bin/env python
import argparse
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Tuple

from openpyxl import load_workbook


def now_beirut() -> str:
    # Asia/Beirut is UTC+2 or +3 depending on DST; we rely on local TZ env to format
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S EET")


def load_ws_header(ws) -> Tuple[List[str], int]:
    rows = list(ws.iter_rows(values_only=True))
    header = None
    start = 0
    for i, r in enumerate(rows):
        if any(c is not None for c in r):
            header = [str(c) if c is not None else "" for c in r]
            start = i + 1
            break
    return header or [], start


def append_row(ws, values: List[str]):
    ws.append(values)


def update_orders(ws, header: List[str], order_id: str, status: str, evidence_dir: str, changelog: List[Dict]):
    idx = {h: i for i, h in enumerate(header)}
    for row in ws.iter_rows(min_row=2):
        cell_val = str(row[idx.get("Order_ID", 0)].value or "")
        if cell_val == order_id:
            def set_field(field: str, value):
                pos = idx.get(field)
                if pos is None:
                    return
                old = row[pos].value
                if old == value:
                    return
                row[pos].value = value
                changelog.append({"Sheet": "ORDERS", "Row_Key": order_id, "Field": field, "Old": old, "New": value})

            set_field("Status", status)
            set_field("Evidence_Artifacts", evidence_dir)
            set_field("Last_Updated_At", now_beirut())
            set_field("Last_Updated_By", "Copilot")
            break


def update_features(ws, header: List[str], feature_ids: List[str], end_to_end: bool, evidence_dir: str, changelog: List[Dict]):
    if not feature_ids:
        return
    idx = {h: i for i, h in enumerate(header)}
    for row in ws.iter_rows(min_row=2):
        fid = str(row[idx.get("Feature_ID", 0)].value or "")
        if fid in feature_ids:
            def set_field(field: str, value):
                pos = idx.get(field)
                if pos is None:
                    return
                old = row[pos].value
                if old == value:
                    return
                row[pos].value = value
                changelog.append({"Sheet": "FULL_STACK_FEATURE_MATRIX", "Row_Key": fid, "Field": field, "Old": old, "New": value})

            set_field("Is_End_to_End_Working", "YES" if end_to_end else "NO")
            set_field("Evidence_Ref", evidence_dir)
            set_field("Last_Updated_At", now_beirut())
            set_field("Last_Updated_By", "Copilot")


def update_tasks(ws, header: List[str], task_ids: List[str], status: str, changelog: List[Dict]):
    if not task_ids:
        return
    idx = {h: i for i, h in enumerate(header)}
    for row in ws.iter_rows(min_row=2):
        tid = str(row[idx.get("Task_ID", 0)].value or "")
        if tid in task_ids:
            def set_field(field: str, value):
                pos = idx.get(field)
                if pos is None:
                    return
                old = row[pos].value
                if old == value:
                    return
                row[pos].value = value
                changelog.append({"Sheet": "EXECUTION_BACKLOG", "Row_Key": tid, "Field": field, "Old": old, "New": value})

            set_field("Status (Open/In Progress/Done)", status)
            set_field("Updated_At", now_beirut())


def append_changelog(ws, header: List[str], items: List[Dict], run_at: str, evidence_dir: str):
    if not items:
        return
    existing = sum(1 for _ in ws.iter_rows(min_row=2, values_only=True) if _)
    for i, item in enumerate(items, start=existing + 1):
        row = [None] * len(header)
        def set_field(name: str, value):
            if name in header:
                row[header.index(name)] = value

        set_field("Change_ID", f"CHG-{i:04d}")
        set_field("Run_At", run_at)
        set_field("Sheet", item.get("Sheet"))
        set_field("Row_Key", item.get("Row_Key"))
        set_field("Field", item.get("Field"))
        set_field("Old_Value", item.get("Old"))
        set_field("New_Value", item.get("New"))
        set_field("Reason", "Gate run update")
        set_field("Evidence_Ref", evidence_dir)
        ws.append(row)


def append_prompt_log(ws, header: List[str], run_at: str, result_status: str, evidence_dir: str):
    row = [None] * len(header)
    def set_field(name: str, value):
        if name in header:
            row[header.index(name)] = value
    existing = sum(1 for _ in ws.iter_rows(min_row=2, values_only=True) if _)
    set_field("Prompt_ID", f"PR-{existing+1:04d}")
    set_field("Run_At", run_at)
    set_field("Prompt_Text", "loop_auto execution")
    set_field("Target_Layer", "loop")
    set_field("Expected_Output", "gate execution + Excel sync")
    set_field("Result_Summary", "Updated via evidence script")
    set_field("Result_Status (Success/Partial/Failed)", "Success" if result_status == "Done" else "Failed")
    set_field("Excel_Updated (YES/NO)", "YES")
    set_field("Notes", evidence_dir)
    ws.append(row)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--excel", required=True)
    ap.add_argument("--order-id", required=True)
    ap.add_argument("--status", required=True)
    ap.add_argument("--evidence-dir", required=True)
    ap.add_argument("--feature-ids", default="")
    ap.add_argument("--task-ids", default="")
    ap.add_argument("--end-to-end", action="store_true")
    args = ap.parse_args()

    path = Path(args.excel).expanduser().resolve()
    wb = load_workbook(path)

    changelog: List[Dict] = []
    run_at = now_beirut()

    # ORDERS
    ws_orders = wb["ORDERS"]
    header_orders, _ = load_ws_header(ws_orders)
    update_orders(ws_orders, header_orders, args.order_id, args.status, args.evidence_dir, changelog)

    # FULL_STACK_FEATURE_MATRIX
    ws_feat = wb["FULL_STACK_FEATURE_MATRIX"]
    header_feat, _ = load_ws_header(ws_feat)
    features = [x for x in args.feature_ids.split(",") if x]
    update_features(ws_feat, header_feat, features, args.end_to_end, args.evidence_dir, changelog)

    # EXECUTION_BACKLOG
    ws_tasks = wb["EXECUTION_BACKLOG"]
    header_tasks, _ = load_ws_header(ws_tasks)
    tasks = [x for x in args.task_ids.split(",") if x]
    update_tasks(ws_tasks, header_tasks, tasks, args.status, changelog)

    # CHANGELOG
    ws_changelog = wb["CHANGELOG"]
    header_changelog, _ = load_ws_header(ws_changelog)
    append_changelog(ws_changelog, header_changelog, changelog, run_at, args.evidence_dir)

    # PROMPT_LOG
    ws_prompt = wb["PROMPT_LOG"]
    header_prompt, _ = load_ws_header(ws_prompt)
    append_prompt_log(ws_prompt, header_prompt, run_at, args.status, args.evidence_dir)

    wb.save(path)


if __name__ == "__main__":
    main()