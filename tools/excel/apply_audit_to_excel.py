#!/usr/bin/env python3
"""
Apply audit results to Excel.

Reads audit.json from a FINAL_AUDIT evidence folder and updates Excel:
- Sets Status="NO-GO" for orders that failed audit
- Updates Last_Updated_At to current Beirut time
- Appends CHANGELOG entries for each modification

Usage:
  python3 tools/excel/apply_audit_to_excel.py --audit-dir docs/evidence/FINAL_AUDIT/<RUN_TS>
"""

import argparse
import json
import sys
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


DEBUG = False


def eprint(*args, **kwargs):
    """Print to stderr."""
    print(*args, file=sys.stderr, **kwargs)


def get_timestamp() -> str:
    """Get current timestamp in Beirut TZ."""
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S %Z")


def load_audit_results(audit_dir: Path):
    """Load audit.json from audit directory."""
    audit_json_path = audit_dir / "audit.json"
    if not audit_json_path.exists():
        eprint(f"ERROR: audit.json not found in {audit_dir}")
        sys.exit(1)

    try:
        with open(audit_json_path, 'r') as f:
            return json.load(f)
    except Exception as e:
        eprint(f"ERROR: Failed to load audit.json: {e}")
        sys.exit(1)


def apply_audit_to_orders(wb, audit_results: dict, audit_evidence_ref: str):
    """Apply audit results to ORDERS sheet."""
    if "ORDERS" not in wb.sheetnames:
        eprint("ERROR: ORDERS sheet not found")
        return False

    ws = wb["ORDERS"]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        eprint("ERROR: ORDERS sheet is empty")
        return False

    header = [str(c) if c is not None else "" for c in rows[0]]
    
    # Find column indices
    try:
        order_id_idx = header.index("Order_ID") + 1
        status_idx = header.index("Status") + 1
        updated_at_idx = header.index("Last_Updated_At") + 1 if "Last_Updated_At" in header else None
    except ValueError as e:
        eprint(f"ERROR: Missing required column: {e}")
        return False

    # Get failed orders from audit
    failed_orders = {
        r["order_id"]: r["errors"]
        for r in audit_results.get("results", [])
        if not r.get("passed", True)
    }

    if not failed_orders:
        print("No failed orders to update.")
        return True

    print(f"Updating {len(failed_orders)} failed orders to NO-GO status...")

    timestamp = get_timestamp()
    changes = []

    # Update each failed order
    for row_idx, row in enumerate(rows[1:], start=2):
        order_id = str(row[order_id_idx - 1]).strip() if row[order_id_idx - 1] else ""
        if order_id in failed_orders:
            old_status = str(row[status_idx - 1]).strip() if row[status_idx - 1] else ""
            
            # Update Status
            ws.cell(row=row_idx, column=status_idx).value = "NO-GO"
            
            # Update Last_Updated_At if column exists
            if updated_at_idx:
                ws.cell(row=row_idx, column=updated_at_idx).value = timestamp
            
            changes.append((order_id, old_status, "NO-GO", failed_orders[order_id]))
            print(f"  {order_id}: {old_status} → NO-GO")

    # Append to CHANGELOG
    if "CHANGELOG" in wb.sheetnames and changes:
        ws_cl = wb["CHANGELOG"]
        changelog_rows = list(ws_cl.iter_rows(values_only=True))
        next_row = len(changelog_rows) + 1

        for order_id, old_status, new_status, errors in changes:
            change_id = f"AUDIT-{timestamp.replace(' ', '-').replace(':', '')}"
            reason = f"Final evidence audit failed: {'; '.join(errors[:2])}"  # First 2 errors
            
            ws_cl.cell(row=next_row, column=1).value = change_id
            ws_cl.cell(row=next_row, column=2).value = timestamp
            ws_cl.cell(row=next_row, column=3).value = "ORDERS"
            ws_cl.cell(row=next_row, column=4).value = order_id
            ws_cl.cell(row=next_row, column=5).value = "Status"
            ws_cl.cell(row=next_row, column=6).value = old_status
            ws_cl.cell(row=next_row, column=7).value = new_status
            ws_cl.cell(row=next_row, column=8).value = reason
            ws_cl.cell(row=next_row, column=9).value = audit_evidence_ref
            next_row += 1

    return True


def main():
    global DEBUG
    ap = argparse.ArgumentParser(description="Apply audit results to Excel")
    ap.add_argument("--audit-dir", required=True,
                    help="Path to FINAL_AUDIT/<RUN_TS> directory with audit.json")
    ap.add_argument("--excel", default="UrbanPoints_CTO_Master_Control_v4.xlsx",
                    help="Path to Excel control file")
    ap.add_argument("--debug", action="store_true", help="Enable debug output")
    args = ap.parse_args()

    DEBUG = args.debug

    # Paths
    audit_dir = Path(args.audit_dir).expanduser().resolve()
    excel_path = Path(args.excel).expanduser().resolve()

    if not audit_dir.exists():
        eprint(f"ERROR: Audit directory not found: {audit_dir}")
        sys.exit(1)

    if not excel_path.exists():
        eprint(f"ERROR: Excel file not found: {excel_path}")
        sys.exit(1)

    # Load audit results
    audit_results = load_audit_results(audit_dir)
    
    if audit_results.get("overall_pass", False):
        print("Audit passed. No Excel updates needed.")
        sys.exit(0)

    # Load Excel
    try:
        wb = load_workbook(excel_path)
    except Exception as e:
        eprint(f"ERROR: Failed to load Excel: {e}")
        sys.exit(1)

    # Apply changes
    audit_evidence_ref = f"excel:{audit_dir.relative_to(excel_path.parent.parent)}/audit.md"
    if not apply_audit_to_orders(wb, audit_results, audit_evidence_ref):
        sys.exit(1)

    # Save Excel
    try:
        wb.save(excel_path)
        print(f"\nExcel updated successfully: {excel_path}")
    except Exception as e:
        eprint(f"ERROR: Failed to save Excel: {e}")
        sys.exit(1)


if __name__ == "__main__":
    main()
