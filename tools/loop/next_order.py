#!/usr/bin/env python
import argparse
import json
import sys
from pathlib import Path
from typing import Dict, List, Optional

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException


PRIORITY_ORDER = {"P0": 0, "P1": 1, "P2": 2}
DEBUG = False


def eprint(*args, **kwargs):
    """Print to stderr."""
    print(*args, file=sys.stderr, **kwargs)


def output_json(ok: bool, **kwargs):
    """Output JSON response to stdout and exit."""
    response = {"ok": ok}
    response.update(kwargs)
    print(json.dumps(response))
    sys.exit(0)


def load_orders(path: Path) -> tuple[List[Dict[str, str]], Optional[str]]:
    """Load orders from Excel. Returns (records, error_message)."""
    # Validate Excel exists
    if not path.exists():
        return [], f"Excel file not found: {path}"

    try:
        wb = load_workbook(path, data_only=True)
    except InvalidFileException as e:
        return [], f"Invalid Excel file: {e}"
    except Exception as e:
        return [], f"Failed to load Excel: {e}"

    # Validate ORDERS sheet exists
    if "ORDERS" not in wb.sheetnames:
        return [], f"Sheet 'ORDERS' not found. Available: {', '.join(wb.sheetnames)}"

    try:
        ws = wb["ORDERS"]
    except Exception as e:
        return [], f"Failed to read ORDERS sheet: {e}"

    # Validate required columns
    rows = list(ws.iter_rows(values_only=True))
    header = None
    for i, r in enumerate(rows):
        if any(c is not None for c in r):
            header = [str(c) if c is not None else "" for c in r]
            start = i + 1
            break

    if header is None:
        return [], "ORDERS sheet is empty or has no header"

    required_cols = {"Order_ID", "Status", "Gate_Command", "Feature_IDs", "Priority", "Depends_On_Orders"}
    missing = required_cols - set(header)
    if missing:
        return [], f"Missing columns in ORDERS: {', '.join(sorted(missing))}"

    if DEBUG:
        eprint(f"[DEBUG] Header columns: {header}")

    records = []
    for r in rows[start:]:
        if all(c is None for c in r):
            continue
        obj: Dict[str, str] = {}
        for h, v in zip(header, r):
            if h:
                obj[h] = v if v is not None else ""
        records.append(obj)

    if DEBUG:
        eprint(f"[DEBUG] Loaded {len(records)} order records")

    return records, None


def split_csv(val: str) -> List[str]:
    if not val:
        return []
    return [x.strip() for x in str(val).split(",") if x and x.strip()]


def pick_next_order(records: List[Dict[str, str]]) -> Optional[Dict[str, str]]:
    """Pick the next executable order by status and dependencies."""
    # Only Open or In Progress
    candidates = [r for r in records if str(r.get("Status", "")).strip() in ("Open", "In Progress")]
    if DEBUG:
        eprint(f"[DEBUG] {len(candidates)} candidates (Open or In Progress)")

    if not candidates:
        return None

    # Build lookup for completion
    status_map = {str(r.get("Order_ID", "")): str(r.get("Status", "")) for r in records}

    def deps_done(rec: Dict[str, str]) -> bool:
        deps = split_csv(rec.get("Depends_On_Orders", ""))
        for d in deps:
            if status_map.get(d) != "Done":
                if DEBUG:
                    eprint(f"[DEBUG] Dependency {d} not Done (status={status_map.get(d)})")
                return False
        return True

    filtered = [r for r in candidates if deps_done(r)]
    if DEBUG:
        eprint(f"[DEBUG] {len(filtered)} candidates after dependency filter")

    if not filtered:
        return None

    def sort_key(rec: Dict[str, str]):
        pr = str(rec.get("Priority", "P2"))
        return (PRIORITY_ORDER.get(pr, 99),)

    filtered.sort(key=sort_key)
    selected = filtered[0]
    if DEBUG:
        eprint(f"[DEBUG] Selected: {selected.get('Order_ID', '?')}")
    return selected


def main():
    global DEBUG
    ap = argparse.ArgumentParser(description="Pick next order from Excel execution queue")
    ap.add_argument("--excel", default="UrbanPoints_CTO_Master_Control_v4.xlsx",
                    help="Path to Excel control file")
    ap.add_argument("--debug", action="store_true", help="Enable debug output to stderr")
    args = ap.parse_args()

    DEBUG = args.debug

    path = Path(args.excel).expanduser().resolve()
    
    if DEBUG:
        eprint(f"[DEBUG] Excel path: {path}")
        eprint(f"[DEBUG] Excel exists: {path.exists()}")

    # Load and validate
    records, error = load_orders(path)
    if error:
        output_json(
            False,
            error=error,
            blockers=[error],
            details=f"Failed to load Excel from {path}"
        )
        return  # unreachable but for clarity

    # No open orders
    rec = pick_next_order(records)
    if not rec:
        output_json(
            False,
            error="no_open_orders",
            blockers=["No open orders with satisfied dependencies"],
            details="All orders are Done or have unsatisfied dependencies"
        )
        return  # unreachable but for clarity

    # Success
    feature_ids = split_csv(rec.get("Feature_IDs", ""))
    task_ids = split_csv(rec.get("Task_IDs", ""))
    output_json(
        True,
        order_id=rec.get("Order_ID", ""),
        gate_script=rec.get("Gate_Script_Path", ""),
        gate_command=rec.get("Gate_Command", ""),
        feature_ids=feature_ids,
        task_ids=task_ids,
        priority=rec.get("Priority", "P2")
    )


if __name__ == "__main__":
    main()